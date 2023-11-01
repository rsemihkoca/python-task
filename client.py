from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Color
from datetime import datetime, timedelta
import argparse
import httpx
import asyncio
import json


API_ENDPOINT = "http://0.0.0.0:8000/process_csv"

def hex_to_rgb(hex_color):
    """Return (red, green, blue) for the color given as #rrggbb."""
    red = int(hex_color[1:3], 16)
    green = int(hex_color[3:5], 16)
    blue = int(hex_color[5:7], 16)

    return f"FF{red:02X}{green:02X}{blue:02X}"

# Function to determine cell background color based on 'hu' value
def get_cell_color(hu):
    today = datetime.today()
    hu = datetime.strptime(hu, '%Y-%m-%d')

    if hu < today - timedelta(days=90):
        return PatternFill(start_color="007500", end_color="007500", fill_type="solid")  # Green
    elif hu < today - timedelta(days=365):
        return PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange
    else:
        return PatternFill(start_color="b30000", end_color="b30000", fill_type="solid")  # Red

async def main(file_path:str, keys:str, colored:str)->None:
    async with httpx.AsyncClient(timeout=100) as client:

        with open(file_path, "rb") as file:
            files = {"file": (file_path, file)}

            # Send a POST request to the server
            response = await client.post(API_ENDPOINT, files=files) #, data=data)

        # Check if the response status code is successful
        if response.status_code == 200:
            # Create an Excel workbook
            wb = Workbook()
            ws = wb.active

            server_data = json.loads(response.json())

            # Sort by 'gruppe' field gruppe can be none
            server_data = sorted(server_data, key=lambda x: (x["gruppe"] is None, x["gruppe"]))

            # Add headers to the Excel sheet
            headers = ['rnr']
            if keys:
                headers.extend(keys)
            ws.append(headers)

            # Add data to the Excel sheet
            for key, item in enumerate(server_data, start=2):
                row = [item['rnr']]
                if keys:
                    row.extend([item[key] if key in item else '' for key in keys])
                ws.append(row)

                if colored:
                    hu = item.get('hu', None)
                    if hu:
                        cell_color = get_cell_color(hu)

                        for cell in ws[key]:
                            cell.fill = cell_color

                if 'labelIds' in headers and 'colorCode' in item and item["colorCode"] is not None and item["colorCode"] != "":
                    color_code = item['colorCode']
                    for cell in ws[key]:
                        current_font = cell.font

                        # Create a new Font object by copying the existing font and changing the color
                        new_font = Font(name=current_font.name, size=current_font.size, bold=current_font.bold, italic=current_font.italic, 
                                        underline=current_font.underline, strike=current_font.strike, color=Color(rgb=hex_to_rgb(color_code)))

                        # Apply the modified font back to the cell
                        cell.font = new_font


            # Save the Excel file with the current date in ISO format
            current_date_iso_formatted = datetime.now().isoformat()[:19].replace(':', '-')
            excel_filename = f'vehicles_{current_date_iso_formatted}.xlsx'
            wb.save(excel_filename)
            print(f"Excel file '{excel_filename}' saved.")
        else:
            print(f"Server returned an error: {response.status_code}")


if __name__ == "__main__":

    parser = argparse.ArgumentParser(description="Client script for sending CSV data to the server.")
    parser.add_argument("-f", "--csv-file", required=True, help="Input CSV file")
    parser.add_argument("-k", "--keys", nargs='+', help="Additional columns to include in the Excel file")
    parser.add_argument("-c", "--colored", action="store_false", help="Color rows based on hu age")
    args = parser.parse_args()

    print("Parsed arguments:")
    print(f"CSV file: {args.csv_file}")
    print(f"Keys: {args.keys}")
    print(f"Colored: {args.colored}")


    asyncio.run(main(args.csv_file, args.keys, args.colored))